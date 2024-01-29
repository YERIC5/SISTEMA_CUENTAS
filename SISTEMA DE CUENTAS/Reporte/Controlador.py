import os
from tkinter import filedialog, messagebox
from Inicio_Sesion.Modelo import Usuario
from Reporte.Modelo import reporte_cuenta, reporte_usuario, insert_reporte

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from datetime import date
import shutil
import openpyxl
from openpyxl.styles import Border, Side, Font

class Reporte:
    def __init__(self, vistaR, username):
        self.vistaR = vistaR
        self.ventana_Reporte = vistaR.masterR
        
        self.user_name = username
        self.id_user, self.rol = Usuario(self.user_name)
        
    def Reporte_Cuenta(self):
        dni = self.vistaR.entry_Cdni.get()
        fecha_Incio = self.vistaR.entry_Cfecha_inicio.get()
        fecha_Final = self.vistaR.entry_Cfecha_final.get()
        
        if not fecha_Incio:
            fecha_Incio = None
            
        if not fecha_Final:
            fecha_Final = None
        
        if not dni:
            dni = None
        
        self.resultado_Cuenta = reporte_cuenta(dni, fecha_Incio, fecha_Final, self.rol, self.id_user)
        
        # Limpiar TreeView antes de agregar nuevos datos
        for record in self.vistaR.tbl_reportCuenta.get_children():
            self.vistaR.tbl_reportCuenta.delete(record)
        
        # Agregar datos al TreeView desde la lista de datos
        for dato in self.resultado_Cuenta:
            self.vistaR.tbl_reportCuenta.insert("", "end", values=dato)
            
    def Reporte_Usuario(self):
        
        if self.rol == "ADMINISTRADOR":
            dni = self.vistaR.entry_Udni.get()
            fecha_Incio = self.vistaR.entry_Ufecha_inicio.get()
            fecha_Final = self.vistaR.entry_Ufecha_final.get()
            
            if not fecha_Incio:
                fecha_Incio = None
                
            if not fecha_Final:
                fecha_Final = None
            
            if not dni:
                dni = None
            
            self.resultado_usuario = reporte_usuario(dni, fecha_Incio, fecha_Final)
            
            # Limpiar TreeView antes de agregar nuevos datos
            for record in self.vistaR.tbl_reportUsuario.get_children():
                self.vistaR.tbl_reportUsuario.delete(record)
            
            # Agregar datos al TreeView desde la lista de datos
            for dato in self.resultado_usuario:
                self.vistaR.tbl_reportUsuario.insert("", "end", values=dato)
        else:
            for frame in [self.vistaR.frameRU_4, self.vistaR.frameRU_2]:
                for child in frame.winfo_children():
                    child.configure(state='disabled')
            self.vistaR.tbl_reportUsuario.state(["disabled"])
            
    def Exportar_Cuenta_Pdf(self):
        try:
            self.tipo_reporte_cuenta = "CUENTA"
            nombre_archivo = "Reporte_Cuenta.pdf"
            
            # Crear un archivo PDF en orientación horizontal
            pdf = canvas.Canvas(nombre_archivo, pagesize=landscape(A4))
            
            # Definir las coordenadas iniciales para la tabla
            x_start = 30
            y_start = 510
            
            # Formato y Estilo de Texto
            letra_tabla = pdf.beginText(x_start, y_start)
            letra_tabla.setFont("Times-Roman", 10)
            pdf.drawText(letra_tabla)
            pdf.setStrokeColor("blue")
            
            # Dibujar encabezados de columna y Definir anchos de celda para cada columna
            column_widths = [25, 80, 100, 130, 80, 170, 100, 140]     
            headers = ["N°", "USUARIO", "CONTRASEÑA", "CUENTA", "DUEÑO", "DESCRIPCION", "LINK", "FECHA CREACION"]
            
            self.Insertar_Datos_Pdf(pdf, column_widths, y_start, x_start, headers, self.resultado_Cuenta)
                
            # Crear el objeto text para el título
            text_titulo = pdf.beginText(30, 550)
            text_titulo.setFont("Helvetica-Bold", 15)
            text_titulo.textLine("REPORTE DE CUENTAS")
            pdf.setFillColorCMYK(0.8,0,0,0.3)
            pdf.drawText(text_titulo)
            
            # Usuario
            
            text_titulo = pdf.beginText(680, 550)
            text_titulo.setFont("Times-Roman", 12)
            text_titulo.textLine(f"Usuario: {self.user_name}")
            pdf.drawText(text_titulo)
            
            # Fecha
            fecha = date.today()
            fecha_creacion = fecha.strftime("%Y-%m-%d")
            text_fecha = pdf.beginText(680, 50)
            text_fecha.setFont("Times-Roman", 12)
            text_fecha.textLine(f"Fecha: {fecha_creacion}")
            pdf.drawText(text_fecha)
            
            # Guardar el PDF
            pdf.save()
            
            # Mostrar el cuadro de diálogo para seleccionar la ubicación y nombre del archivo
            ruta_guardado = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("Archivos PDF", "*.pdf")])
            
            # Copiar el archivo a la ubicación seleccionada por el usuario
            shutil.copy2(nombre_archivo, ruta_guardado)
            
            # Abrir el archivo PDF con la aplicación predeterminada del sistema
            os.startfile(ruta_guardado)
            
            self.inserta_reporte(self.tipo_reporte_cuenta, nombre_archivo, ruta_guardado)
        
        except Exception as e:
            messagebox.showwarning("Error", "No se puedo descargar")
            
    def Exportar_Usuario_Pdf(self):
        try:
            self.tipo_reporte_usuario = "USUARIO"
            nombre_archivo = "Reporte_Usuario.pdf"
            
            # Crear un archivo PDF en orientación horizontal
            pdf = canvas.Canvas(nombre_archivo, pagesize=landscape(A4))
            
            # Definir las coordenadas iniciales para la tabla
            x_start = 30
            y_start = 510
            
            # Formato y Estilo de Texto
            letra_tabla = pdf.beginText(x_start, y_start)
            letra_tabla.setFont("Times-Roman", 10)
            pdf.drawText(letra_tabla)
            pdf.setStrokeColor("blue")
            
            # Dibujar encabezados de columna y Definir anchos de celda para cada columna
            column_widths = [20, 50, 170, 150, 60, 90, 40, 70, 60, 70]     
            headers = ["N°", "DNI", "NOMBRE COMPLETO", "EMAIL", "TELEFONO", "ROL", "EDAD", "NACIMIENTO", "USUARIO", "CONTRASEÑA"]
            
            
            self.Insertar_Datos_Pdf(pdf, column_widths, y_start, x_start, headers, self.resultado_usuario)
                
            # Crear el objeto text para el título
            text_titulo = pdf.beginText(30, 550)
            text_titulo.setFont("Helvetica-Bold", 15)
            text_titulo.textLine("REPORTE DE USUARIOS")
            pdf.setFillColorCMYK(0.8,0,0,0.3)
            pdf.drawText(text_titulo)
            
            # Usuario
            text_titulo = pdf.beginText(680, 550)
            text_titulo.setFont("Times-Roman", 12)
            text_titulo.textLine(f"Usuario: {self.user_name}")
            pdf.drawText(text_titulo)
            
            # Fecha
            fecha = date.today()
            fecha_creacion = fecha.strftime("%Y-%m-%d")
            text_fecha = pdf.beginText(680, 50)
            text_fecha.setFont("Times-Roman", 12)
            text_fecha.textLine(f"Fecha: {fecha_creacion}")
            pdf.drawText(text_fecha)
            
            # Guardar el PDF
            pdf.save()
            
            # Mostrar el cuadro de diálogo para seleccionar la ubicación y nombre del archivo
            ruta_guardado = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("Archivos PDF", "*.pdf")])
            
            # Copiar el archivo a la ubicación seleccionada por el usuario
            shutil.copy2(nombre_archivo, ruta_guardado)
            
            # Abrir el archivo PDF con la aplicación predeterminada del sistema
            os.startfile(ruta_guardado)
            
            self.inserta_reporte(self.tipo_reporte_usuario, nombre_archivo, ruta_guardado)
        
        except Exception as e:
            messagebox.showwarning("Error", "No se puedo descargar")
            
    def Insertar_Datos_Pdf(self, pdf, column_widths, y_start, x_start, headers, resultados):
        for i, (header, width) in enumerate(zip(headers, column_widths)):
                pdf.line(30, 505, 810, 505)
                pdf.drawString(x_start + sum(column_widths[:i]), y_start, header)
            
        # Dibujar los datos en la tabla
        cell_height = 20
        for row_num, row in enumerate(resultados, start=1):
            for col_num, (cell_value, width) in enumerate(zip(row, column_widths)):
                pdf.drawString(x_start + sum(column_widths[:col_num]), y_start - row_num * cell_height, str(cell_value))
    
    def Exportar_Cuenta_Excel(self):
        try:
            nombre_archivo = "Reporte_Cuentas.xlsx"
            # Crear un nuevo libro de trabajo (workbook)
            libro = openpyxl.Workbook()
            
            # Seleccionar la hoja de trabajo activa (por defecto es la primera)
            hoja = libro.active
            
            self.estilos_xlsx()             
            
            # Datos que deseas insertar (puedes adaptar esta lista según tus necesidades)
            hoja.title = "REPORTE DE CUENTAS"
            titulo = ["REPORTE DE CUENTAS"]
            hoja.append(titulo)
            user = [f"USUARIO: {self.user_name}"]
            hoja.append(user)
            
            # Fecha
            fecha = date.today()
            fecha_creacion = fecha.strftime("%Y-%m-%d")
            fecha = [f"FECHA: {fecha_creacion}"]
            hoja.append(fecha)
            nombre_column = ["N°", "USUARIO", "CONTRASEÑA", "CUENTA", "DUEÑO", "DESCRIPCION", "LINK", "FECHA DE CREACION"]
            hoja.append(nombre_column)
            
            hoja.merge_cells('A1:H1')
            hoja.merge_cells('A2:H2')
            hoja.merge_cells('A3:H3')
            
            # Aplicar estilos a filas específicas
            self.aplicar_estilo_fila(hoja, 1, self.estilo_titulo)
            self.aplicar_estilo_fila(hoja, 2, self.estilo_user_fecha)
            self.aplicar_estilo_fila(hoja, 3, self.estilo_user_fecha)
            self.aplicar_estilo_fila(hoja, 4, self.estilo_columnas)
            
            # Insertar los datos en la hoja de trabajo
            for fila in self.resultado_Cuenta:
                hoja.append(fila)
                
            # Crear un borde
            borde = Border(
                left=Side(border_style='medium', color='000000'),
                right=Side(border_style='medium', color='000000'),
                top=Side(border_style='medium', color='000000'),
                bottom=Side(border_style='medium', color='000000')
            )
            
            # Aplicar estilos a filas específicas
            self.aplicar_estilo_fila(hoja, 1, self.estilo_titulo)
            self.aplicar_estilo_fila(hoja, 2, self.estilo_user_fecha)
            
            # Aplicar el borde a todas las celdas en la hoja
            for fila in hoja.iter_rows(min_row=1, max_col=hoja.max_column, max_row=hoja.max_row):
                for celda in fila:
                    celda.border = borde
                    
            # Ajustar el tamaño de las columnas
            for i, titulo in enumerate(nombre_column, start=1):
                longitud_maxima = max(len(str(titulo)), *[len(str(fila[i-1])) for fila in self.resultado_Cuenta])
                hoja.column_dimensions[chr(64+i)].width = longitud_maxima + 2
                
            # Guardar el libro de trabajo en un archivo Excel
            libro.save(nombre_archivo)
            
            # Mostrar el cuadro de diálogo para seleccionar la ubicación y nombre del archivo
            ruta_guardado = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos XLSX", "*.xlsx")])
            
            # Copiar el archivo a la ubicación seleccionada por el usuario
            shutil.copy2(nombre_archivo, ruta_guardado)
            
            # Abrir el archivo PDF con la aplicación predeterminada del sistema
            os.startfile(ruta_guardado)
            self.inserta_reporte(self.tipo_reporte_cuenta, nombre_archivo, ruta_guardado)
            
        except Exception as e:
            messagebox.showwarning("Error", "No se puedo descargar")
        
    def Exportar_Usuario_Excel(self):
        try:
            nombre_archivo = "Reporte_Usuarios.xlsx"
            # Crear un nuevo libro de trabajo (workbook)
            libro = openpyxl.Workbook()
            
            # Seleccionar la hoja de trabajo activa (por defecto es la primera)
            hoja = libro.active    
            
            # Datos que deseas insertar (puedes adaptar esta lista según tus necesidades)
            hoja.title = "REPORTE DE USUARIOS"
            titulo = ["REPORTE DE USUARIOS"]
            hoja.append(titulo)
            user = [f"USUARIO: {self.user_name}"]
            hoja.append(user)
            
            # Fecha
            fecha = date.today()
            fecha_creacion = fecha.strftime("%Y-%m-%d")
            fecha = [f"FECHA: {fecha_creacion}"]
            hoja.append(fecha)
            nombre_column = ["N°", "DNI", "NOMBRE COMPLETO", "EMAIL", "TELEFONO", "ROL", "EDAD", "FECHA DE NACIMIENTO", "USUARIO", "CONTRASEÑA"]
            hoja.append(nombre_column)
            
            hoja.merge_cells('A1:J1')
            hoja.merge_cells('A2:J2')
            hoja.merge_cells('A3:J3')
            
            # Aplicar estilos a filas específicas
            self.aplicar_estilo_fila(hoja, 1, self.estilo_titulo)
            self.aplicar_estilo_fila(hoja, 2, self.estilo_user_fecha)
            self.aplicar_estilo_fila(hoja, 3, self.estilo_user_fecha)
            self.aplicar_estilo_fila(hoja, 4, self.estilo_columnas)
            
            # Insertar los datos en la hoja de trabajo
            for fila in self.resultado_usuario:
                hoja.append(fila)
                
            # Crear un borde
            borde = Border(
                left=Side(border_style='medium', color='000000'),
                right=Side(border_style='medium', color='000000'),
                top=Side(border_style='medium', color='000000'),
                bottom=Side(border_style='medium', color='000000')
            )
            
            # Aplicar estilos a filas específicas
            self.aplicar_estilo_fila(hoja, 1, self.estilo_titulo)
            self.aplicar_estilo_fila(hoja, 2, self.estilo_user_fecha)
            
            # Aplicar el borde a todas las celdas en la hoja
            for fila in hoja.iter_rows(min_row=1, max_col=hoja.max_column, max_row=hoja.max_row):
                for celda in fila:
                    celda.border = borde
                    
            # Ajustar el tamaño de las columnas
            for i, titulo in enumerate(nombre_column, start=1):
                longitud_maxima = max(len(str(titulo)), *[len(str(fila[i-1])) for fila in self.resultado_usuario])
                hoja.column_dimensions[chr(64+i)].width = longitud_maxima + 2
                
            # Guardar el libro de trabajo en un archivo Excel
            libro.save(nombre_archivo)
            
            # Mostrar el cuadro de diálogo para seleccionar la ubicación y nombre del archivo
            ruta_guardado = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos XLSX", "*.xlsx")])
            
            # Copiar el archivo a la ubicación seleccionada por el usuario
            shutil.copy2(nombre_archivo, ruta_guardado)
            
            # Abrir el archivo PDF con la aplicación predeterminada del sistema
            os.startfile(ruta_guardado)
            
            self.inserta_reporte(self.tipo_reporte_usuario, nombre_archivo, ruta_guardado)
            
        except Exception as e:
            messagebox.showwarning("Error", "No se puedo descargar")
        
    def aplicar_estilo_fila(self, hoja, numero_fila, estilo):
        for columna in hoja.iter_cols(min_row=numero_fila, max_row=numero_fila):
            for celda in columna:
                celda.font = estilo
                
    def estilos_xlsx(self):
        # Crear un estilo de fuente personalizado
            self.estilo_titulo = Font(
                name='Arial',
                size=15,
                bold=True,
                italic=False,
                color='538DD5'  # Color en formato hexadecimal
            )
            
            self.estilo_user_fecha = Font(
                name='Arial',
                size=11,
                bold=True,
                italic=False,
                color='538DD5'  # Color en formato hexadecimal
            )
            
            self.estilo_columnas = Font(
                name='Arial',
                size=10,
                bold=True,
                italic=False,
                color='C0504D'  # Color en formato hexadecimal
            ) 
            
    def inserta_reporte(self, tipo_reporte, nombre_document, ruta_archivo):        
        documento = None
        if nombre_document.lower().endswith(".pdf"):
            # Leer el archivo en formato binario
            with open(ruta_archivo, 'rb') as archivo:
                documento = archivo.read()
        if nombre_document.lower().endswith(".xlsx"):
            # Leer el archivo en formato binario
            with open(ruta_archivo, 'rb') as archivo:
                documento = archivo.read()
        
        insert_reporte(tipo_reporte, nombre_document, self.id_user, documento)
